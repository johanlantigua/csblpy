"""Python version of jumpplatematlab/jump_plate_executable.m."""

from __future__ import annotations

import inspect
import math
import sys
import time
import tkinter as tk
from collections import namedtuple
from ctypes import windll
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from PIL import Image, ImageTk


APP_NAME = "CSBL Jump Plate"
FAIRFIELD_RED = "#E51937"
FAIRFIELD_BLACK = "#231F20"
FAIRFIELD_GRAY = "#BFC0C2"
APP_BG = "#F4F5F7"
PANEL_BG = "#FFFFFF"
TEXT_DARK = "#171717"
SUCCESS_GREEN = "#1B8A4A"
ERROR_RED = "#C8102E"


def resource_path(*parts: str) -> Path:
    if hasattr(sys, "_MEIPASS"):
        return Path(getattr(sys, "_MEIPASS")).joinpath(*parts)
    return Path(__file__).resolve().parent.parent.joinpath(*parts)


@dataclass
class Sample:
    time_s: float
    jump_plate_v: float
    laser_v: float = math.nan


class ArduinoVoltageReader:
    """Read A0 through pyfirmata. The board must have StandardFirmata uploaded."""

    def __init__(self, port: str) -> None:
        self._patch_pyfirmata_for_python_312()
        try:
            import pyfirmata
            from pyfirmata import util
        except ImportError as exc:
            raise RuntimeError(
                "Arduino support requires pyfirmata. Install it with "
                "`python -m pip install pyfirmata pyserial`."
            ) from exc

        self._board = pyfirmata.Arduino(port)
        self._iterator = util.Iterator(self._board)
        self._iterator.start()
        self._pin = self._board.analog[0]
        self._pin.enable_reporting()
        self._wait_for_analog_data()

    def _patch_pyfirmata_for_python_312(self) -> None:
        if hasattr(inspect, "getargspec"):
            return
        ArgSpec = namedtuple("ArgSpec", "args varargs keywords defaults")

        def getargspec(func):
            spec = inspect.getfullargspec(func)
            return ArgSpec(spec.args, spec.varargs, spec.varkw, spec.defaults)

        inspect.getargspec = getargspec

    def read_voltage_a0(self) -> float:
        value = self._pin.read()
        if value is None:
            return math.nan
        return float(value) * 5.0

    def _wait_for_analog_data(self) -> None:
        deadline = time.monotonic() + 3.0
        while time.monotonic() < deadline:
            value = self._pin.read()
            if value is not None:
                return
            time.sleep(0.05)
        self.close()
        raise RuntimeError(
            "The COM port opened, but no analog data came back from A0.\n\n"
            "Most likely StandardFirmata is not uploaded to this Arduino, or the jump plate is not wired to A0."
        )

    def close(self) -> None:
        self._board.exit()


class JumpPlateApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1280x760")
        self.minsize(960, 570)
        self.configure(bg=APP_BG)
        self._set_windows_identity()
        self._set_app_icon()

        self.arduino: ArduinoVoltageReader | None = None
        self.total_time = 5.0
        self.sample_rate = math.nan
        self.raw_data: list[Sample] = []
        self.time_point_1: float | str = "NaN"
        self.time_point_2: float | str = "NaN"
        self.measured_time: float | str = "NaN"
        self.jump_height: float | str = "NaN"
        self._start_monotonic = 0.0
        self._collecting = False
        self._selecting_point: int | None = None
        self._images: dict[str, ImageTk.PhotoImage] = {}

        self._build_header()
        self._build_plot_area()
        self._build_controls()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _set_windows_identity(self) -> None:
        if sys.platform != "win32":
            return
        try:
            windll.shell32.SetCurrentProcessExplicitAppUserModelID("Fairfield.CSBL.JumpPlate")
        except Exception:
            pass

    def _set_app_icon(self) -> None:
        for icon_path in (
            resource_path("jumpplatepy", "assets", "lablogo.ico"),
            Path(__file__).resolve().parent / "assets" / "lablogo.ico",
        ):
            if icon_path.exists():
                try:
                    self.iconbitmap(default=str(icon_path))
                except tk.TclError:
                    pass
                return

    def _build_header(self) -> None:
        header = tk.Frame(self, bg=FAIRFIELD_BLACK)
        header.place(relx=0, rely=0, relwidth=1, relheight=0.145)
        tk.Frame(self, bg=FAIRFIELD_RED).place(relx=0, rely=0.145, relwidth=1, height=5)

        lab_logo = self._load_image(resource_path("pics", "lablogo.png"), (92, 82))
        if lab_logo:
            tk.Label(header, image=lab_logo, bg=FAIRFIELD_BLACK).place(
                relx=0.022, rely=0.1, relwidth=0.08, relheight=0.78
            )

        title_block = tk.Frame(header, bg=FAIRFIELD_BLACK)
        title_block.place(relx=0.12, rely=0.16, relwidth=0.62, relheight=0.68)
        tk.Label(
            title_block,
            text=APP_NAME,
            bg=FAIRFIELD_BLACK,
            fg="white",
            font=("Segoe UI", 28, "bold"),
            anchor="w",
        ).pack(fill="x")
        tk.Label(
            title_block,
            text="Community-Situated Biomechanics Lab at Fairfield University",
            bg=FAIRFIELD_BLACK,
            fg=FAIRFIELD_GRAY,
            font=("Segoe UI", 12),
            anchor="w",
        ).pack(fill="x")

        fairfield_logo = self._load_image(resource_path("pics", "fairfieldlogo.png"), (94, 82))
        if fairfield_logo:
            tk.Label(header, image=fairfield_logo, bg=FAIRFIELD_BLACK).place(
                relx=0.895, rely=0.1, relwidth=0.075, relheight=0.78
            )

    def _load_image(self, path: Path, max_size: tuple[int, int]) -> ImageTk.PhotoImage | None:
        if not path.exists():
            return None
        image = Image.open(path).convert("RGBA")
        image.thumbnail(max_size, Image.LANCZOS)
        photo = ImageTk.PhotoImage(image)
        self._images[str(path)] = photo
        return photo

    def _build_plot_area(self) -> None:
        self.plot = tk.Canvas(self, bg=PANEL_BG, highlightthickness=1, highlightbackground=FAIRFIELD_GRAY)
        self.plot.place(relx=0.057, rely=0.34, relwidth=0.742, relheight=0.58)
        self.plot.bind("<Button-1>", self._plot_clicked)
        self.plot.bind("<Configure>", lambda _event: self._redraw_plot())

    def _build_controls(self) -> None:
        self.initialize_button = tk.Button(self, text="Initialize Arduino", command=self.initialize_arduino)
        self._style_button(self.initialize_button, primary=True)
        self.initialize_button.place(relx=0.845, rely=0.285, relwidth=0.116, relheight=0.065)

        port_panel = self._panel("Arduino Com Port", 0.84, 0.365, 0.128, 0.065)
        self.port_edit = tk.Entry(port_panel, justify="center")
        self._style_entry(self.port_edit)
        self.port_edit.insert(0, "COM4")
        self.port_edit.place(relx=0.08, rely=0.22, relwidth=0.84, relheight=0.7)

        self.start_button = tk.Button(self, text="Start Data Collection", command=self.start_data_collection)
        self._style_button(self.start_button, primary=True)
        self.start_button.place(relx=0.04, rely=0.18, relwidth=0.32, relheight=0.09)

        duration_panel = self._panel("Sampling Duration", 0.395, 0.18, 0.205, 0.09)
        self.sampling_dur_edit = tk.Entry(duration_panel, justify="center")
        self._style_entry(self.sampling_dur_edit)
        self.sampling_dur_edit.insert(0, "5")
        self.sampling_dur_edit.place(relx=0.05, rely=0.15, relwidth=0.88, relheight=0.78)

        player_panel = self._panel("Enter Player Number", 0.625, 0.18, 0.18, 0.09)
        self.player_number_edit = tk.Entry(player_panel, justify="center")
        self._style_entry(self.player_number_edit)
        self.player_number_edit.insert(0, "Player Number")
        self.player_number_edit.place(relx=0.06, rely=0.15, relwidth=0.88, relheight=0.78)

        jump_panel = self._panel("Vertical Jump (in)", 0.825, 0.18, 0.159, 0.09)
        self.jump_height_text = tk.Label(jump_panel, text="No Data", bg=PANEL_BG, fg=TEXT_DARK)
        self.jump_height_text.place(relx=0.03, rely=0.15, relwidth=0.92, relheight=0.74)

        self.export_data_button = tk.Button(self, text="Export Data", command=self.export_data)
        self._style_button(self.export_data_button)
        self.export_data_button.place(relx=0.82, rely=0.445, relwidth=0.15, relheight=0.065)

        self.first_time_point_button = tk.Button(
            self, text="Select Take Off", command=lambda: self._begin_selecting_point(1)
        )
        self._style_button(self.first_time_point_button)
        self.first_time_point_button.place(relx=0.028, rely=0.285, relwidth=0.117, relheight=0.055)

        time_1_panel = self._panel("Take Off Point", 0.164, 0.285, 0.159, 0.055)
        self.time_point_1_text = tk.Label(time_1_panel, text="No Data", bg=PANEL_BG, fg=TEXT_DARK)
        self.time_point_1_text.place(relx=0.02, rely=0.15, relwidth=0.88, relheight=0.78)

        self.second_time_point_button = tk.Button(
            self, text="Select Landing", command=lambda: self._begin_selecting_point(2)
        )
        self._style_button(self.second_time_point_button)
        self.second_time_point_button.place(relx=0.338, rely=0.285, relwidth=0.117, relheight=0.055)

        time_2_panel = self._panel("Landing Point", 0.472, 0.285, 0.159, 0.055)
        self.time_point_2_text = tk.Label(time_2_panel, text="No Data", bg=PANEL_BG, fg=TEXT_DARK)
        self.time_point_2_text.place(relx=0.05, rely=0.12, relwidth=0.84, relheight=0.78)

        hang_panel = self._panel("Hang Time", 0.647, 0.285, 0.159, 0.055)
        self.measured_time_text = tk.Label(hang_panel, text="No Data", bg=PANEL_BG, fg=TEXT_DARK)
        self.measured_time_text.place(relx=0.03, rely=0.15, relwidth=0.92, relheight=0.74)

    def _panel(self, title: str, relx: float, rely: float, relwidth: float, relheight: float) -> tk.LabelFrame:
        panel = tk.LabelFrame(
            self,
            text=title,
            bg=PANEL_BG,
            fg=FAIRFIELD_BLACK,
            labelanchor="n",
            font=("Segoe UI", 9, "bold"),
            bd=1,
            relief="solid",
        )
        panel.place(relx=relx, rely=rely, relwidth=relwidth, relheight=relheight)
        return panel

    def _style_button(self, button: tk.Button, primary: bool = False) -> None:
        bg = FAIRFIELD_RED if primary else FAIRFIELD_BLACK
        active_bg = "#B90F28" if primary else "#3A3435"
        button.configure(
            bg=bg,
            fg="white",
            activebackground=active_bg,
            activeforeground="white",
            bd=0,
            relief="flat",
            cursor="hand2",
            font=("Segoe UI", 10, "bold"),
            highlightthickness=0,
            padx=6,
            pady=4,
        )

    def _style_entry(self, entry: tk.Entry) -> None:
        entry.configure(
            bg="#FFFFFF",
            fg=TEXT_DARK,
            insertbackground=FAIRFIELD_RED,
            relief="flat",
            font=("Segoe UI", 10),
        )

    def initialize_arduino(self) -> None:
        port = self.port_edit.get().strip()
        try:
            self.arduino = ArduinoVoltageReader(port)
        except Exception as exc:
            self.port_edit.configure(bg="#FFD6DE", fg=ERROR_RED)
            messagebox.showerror("Arduino initialization failed", str(exc))
            return
        self.port_edit.configure(bg="#DDF3E6", fg=SUCCESS_GREEN)

    def start_data_collection(self) -> None:
        if self._collecting:
            return
        try:
            self.total_time = float(self.sampling_dur_edit.get())
        except ValueError:
            messagebox.showerror("Invalid duration", "Sampling Duration must be a number.")
            return
        if self.total_time <= 0:
            messagebox.showerror("Invalid duration", "Sampling Duration must be greater than zero.")
            return
        if self.arduino is None:
            messagebox.showerror("Arduino not initialized", "Click Initialize Arduino before collecting data.")
            return

        self.raw_data = []
        self.time_point_1 = "NaN"
        self.time_point_2 = "NaN"
        self.measured_time = "NaN"
        self.jump_height = "NaN"
        self._selecting_point = None
        self._set_status_label(self.time_point_1_text, "No Data", ERROR_RED)
        self._set_status_label(self.time_point_2_text, "No Data", ERROR_RED)
        self._set_status_label(self.measured_time_text, "No Data", ERROR_RED)
        self._set_status_label(self.jump_height_text, "No Data", ERROR_RED)
        self._collecting = True
        self._start_monotonic = time.monotonic()
        self._redraw_plot()
        self._collect_next_sample()

    def _collect_next_sample(self) -> None:
        elapsed = time.monotonic() - self._start_monotonic
        if elapsed >= self.total_time:
            self._collecting = False
            matlab_i = len(self.raw_data) + 1
            self.sample_rate = matlab_i / self.total_time
            return

        voltage = self.arduino.read_voltage_a0() if self.arduino else math.nan
        self.raw_data.append(Sample(elapsed, voltage))
        self._redraw_plot()
        self.after(1, self._collect_next_sample)

    def _begin_selecting_point(self, point_number: int) -> None:
        if not self.raw_data:
            messagebox.showwarning("No data", "Collect data before selecting points.")
            return
        self._selecting_point = point_number
        self.plot.configure(cursor="crosshair")

    def _plot_clicked(self, event: tk.Event) -> None:
        if self._selecting_point is None or not self.raw_data:
            return
        clicked_time = self._canvas_to_time(event.x)
        out_index = self._first_sample_after(clicked_time)

        if self._selecting_point == 1:
            self.time_point_1 = clicked_time
            self._set_status_label(self.time_point_1_text, self._matlab_num(clicked_time), SUCCESS_GREEN)
        else:
            self.time_point_2 = clicked_time
            if isinstance(self.time_point_1, float):
                self.measured_time = self.time_point_2 - self.time_point_1
                self.jump_height = math.floor((9.8 * self.measured_time**2) / 8 * 3.38 * 12 * 10) / 10
            else:
                self.measured_time = "NaN"
                self.jump_height = "NaN"
            self._set_status_label(self.time_point_2_text, self._matlab_num(clicked_time), SUCCESS_GREEN)
            self._set_status_label(self.measured_time_text, self._matlab_num(self.measured_time), SUCCESS_GREEN)
            self._set_status_label(self.jump_height_text, self._matlab_num(self.jump_height), SUCCESS_GREEN)

        self._selecting_point = None
        self.plot.configure(cursor="")
        self._redraw_plot()

    def _first_sample_after(self, time_s: float) -> int:
        for index, sample in enumerate(self.raw_data):
            if time_s < sample.time_s:
                return index
        return len(self.raw_data) - 1

    def _set_status_label(self, label: tk.Label, text: str, color: str) -> None:
        label.configure(text=text, bg=color, fg="white", font=("Segoe UI", 10, "bold"))

    def _matlab_num(self, value: float | str) -> str:
        if isinstance(value, str):
            return value
        return f"{value:.15g}"

    def export_data(self) -> None:
        if not self.raw_data:
            messagebox.showwarning("No data", "There is no raw data to export.")
            return
        export_dir = filedialog.askdirectory(title="Choose export folder")
        if not export_dir:
            return

        output_path = Path(export_dir) / f"vertical_{self.player_number_edit.get()}.xlsx"
        now = datetime.now()
        test_info = [
            "Player Number",
            "year",
            "month",
            "day",
            "hour",
            "minute",
            "second",
            "Sample Duration",
            "Sample Rate",
            "Take off Time",
            "Landing Time",
            "Time Measured",
            "Jump Height (in)",
        ]
        test_results = [
            self.player_number_edit.get(),
            str(now.year),
            str(now.month),
            str(now.day),
            str(now.hour),
            str(now.minute),
            str(now.second),
            self.total_time,
            self.sample_rate,
            self.time_point_1,
            self.time_point_2,
            self.measured_time,
            self.jump_height,
        ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        for row, (label, value) in enumerate(zip(test_info, test_results), start=1):
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=value)

        sheet.cell(row=2, column=5, value="Time (s)")
        sheet.cell(row=2, column=6, value="Signal Jump Plate (V)")
        sheet.cell(row=2, column=7, value="Signal Laser (V)")
        for row, sample in enumerate(self.raw_data, start=3):
            sheet.cell(row=row, column=5, value=sample.time_s)
            sheet.cell(row=row, column=6, value=sample.jump_plate_v)
            sheet.cell(row=row, column=7, value=sample.laser_v)

        workbook.save(output_path)
        messagebox.showinfo("Export Data", f"Saved {output_path}")

    def _redraw_plot(self) -> None:
        self.plot.delete("all")
        width = max(self.plot.winfo_width(), 2)
        height = max(self.plot.winfo_height(), 2)
        margin_left, margin_right, margin_top, margin_bottom = 52, 18, 18, 38
        x0, y0 = margin_left, height - margin_bottom
        x1, y1 = width - margin_right, margin_top

        self.plot.create_rectangle(x0, y1, x1, y0, outline=FAIRFIELD_BLACK, fill="white", width=2)
        self.plot.create_text((x0 + x1) / 2, height - 12, text="Time (s)", fill=TEXT_DARK)
        self.plot.create_text(16, (y0 + y1) / 2, text="Signal Jump Plate (V)", angle=90, fill=TEXT_DARK)

        for frac in (0, 0.25, 0.5, 0.75, 1):
            x = x0 + frac * (x1 - x0)
            y = y0 - frac * (y0 - y1)
            self.plot.create_line(x, y0, x, y1, fill="#ECEFF1")
            self.plot.create_line(x0, y, x1, y, fill="#ECEFF1")
            self.plot.create_text(x, y0 + 14, text=self._matlab_num(frac * self.total_time), fill=TEXT_DARK)
            self.plot.create_text(x0 - 18, y, text=self._matlab_num(frac * 5), fill=TEXT_DARK)

        if len(self.raw_data) == 1:
            x, y = self._to_canvas(self.raw_data[0], x0, y0, x1, y1)
            self.plot.create_oval(x - 3, y - 3, x + 3, y + 3, fill=FAIRFIELD_BLACK, outline=FAIRFIELD_BLACK)
        elif len(self.raw_data) > 1:
            points: list[float] = []
            for sample in self.raw_data:
                points.extend(self._to_canvas(sample, x0, y0, x1, y1))
            self.plot.create_line(*points, fill=FAIRFIELD_RED, width=2)

        self._draw_selected_point(self.time_point_1, "Taking off", x0, y0, x1, y1)
        self._draw_selected_point(self.time_point_2, "Landing", x0, y0, x1, y1)

    def _draw_selected_point(
        self, point: float | str, label: str, x0: int, y0: int, x1: int, y1: int
    ) -> None:
        if not isinstance(point, float) or not self.raw_data:
            return
        sample = self.raw_data[self._first_sample_after(point)]
        x, y = self._to_canvas(sample, x0, y0, x1, y1)
        self.plot.create_oval(x - 5, y - 5, x + 5, y + 5, fill=FAIRFIELD_BLACK, outline=FAIRFIELD_RED, width=2)
        offset = -55 if label == "Taking off" else 45
        self.plot.create_text(x + offset, y - 10, text=label, fill=FAIRFIELD_BLACK, font=("Segoe UI", 10, "bold"))

    def _to_canvas(self, sample: Sample, x0: int, y0: int, x1: int, y1: int) -> tuple[float, float]:
        x = x0 + (sample.time_s / max(self.total_time, 1e-9)) * (x1 - x0)
        voltage = 0.0 if math.isnan(sample.jump_plate_v) else max(0.0, min(5.0, sample.jump_plate_v))
        y = y0 - (voltage / 5.0) * (y0 - y1)
        return x, y

    def _canvas_to_time(self, x: int) -> float:
        width = max(self.plot.winfo_width(), 2)
        x0, x1 = 52, width - 18
        frac = (x - x0) / max(x1 - x0, 1)
        return max(0.0, min(self.total_time, frac * self.total_time))

    def _on_close(self) -> None:
        if self.arduino is not None:
            self.arduino.close()
        self.destroy()


def main() -> int:
    app = JumpPlateApp()
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
