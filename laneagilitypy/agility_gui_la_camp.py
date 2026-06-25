"""Python version of the MATLAB GUIDE app in matlabcode/agility_GUI_LA_camp.m.

The GUI intentionally mirrors the MATLAB layout and workflow:
initialize an Arduino, record A1 voltage for a set duration, click two time
points on the plot, calculate the elapsed time, and export the results.
"""

from __future__ import annotations

import math
import sys
import time
import tkinter as tk
from ctypes import windll
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import messagebox

from openpyxl import Workbook
from PIL import Image, ImageTk


MATLAB_EXPORT_DIR = Path(
    r"C:\Users\John\Dropbox\4th Family\LA Skills Acadamy\evaluation folder\agility"
)

APP_NAME = "CSBL Agility"
FAIRFIELD_RED = "#E51937"
FAIRFIELD_BLACK = "#231F20"
FAIRFIELD_GRAY = "#BFC0C2"
APP_BG = "#F4F5F7"
PANEL_BG = "#FFFFFF"
TEXT_DARK = "#171717"
SUCCESS_GREEN = "#1B8A4A"
ERROR_RED = "#C8102E"


def resource_path(*parts: str) -> Path:
    """Return a path that works both from source and from a PyInstaller bundle."""
    if hasattr(sys, "_MEIPASS"):
        return Path(getattr(sys, "_MEIPASS")).joinpath(*parts)
    return Path(__file__).resolve().parent.parent.joinpath(*parts)


@dataclass
class Sample:
    time_s: float
    voltage: float


class ArduinoVoltageReader:
    """Read A1 through pyfirmata when the board has StandardFirmata loaded."""

    def __init__(self, port: str) -> None:
        try:
            self._patch_pyfirmata_for_python_312()
            import pyfirmata
            from pyfirmata import util
        except ImportError as exc:
            raise RuntimeError(
                "Arduino support requires pyfirmata. Install it with "
                "`python -m pip install pyfirmata pyserial`."
            ) from exc

        self._board = pyfirmata.Arduino(port)
        self._iterator = util.Iterator(self._board)
        self._iterator.start()
        self._pin = self._board.analog[1]
        self._pin.enable_reporting()
        time.sleep(0.5)

    def _patch_pyfirmata_for_python_312(self) -> None:
        import inspect
        from collections import namedtuple

        if hasattr(inspect, "getargspec"):
            return

        ArgSpec = namedtuple("ArgSpec", "args varargs keywords defaults")

        def getargspec(func):
            spec = inspect.getfullargspec(func)
            return ArgSpec(spec.args, spec.varargs, spec.varkw, spec.defaults)

        inspect.getargspec = getargspec

    def read_voltage_a1(self) -> float:
        value = self._pin.read()
        if value is None:
            return math.nan
        return float(value) * 5.0

    def close(self) -> None:
        self._board.exit()


class AgilityGui(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1280x760")
        self.minsize(960, 570)
        self.configure(bg=APP_BG)
        self._set_windows_identity()
        self._set_app_icon()

        self.arduino: ArduinoVoltageReader | None = None
        self.total_time = 10.0
        self.sample_rate = math.nan
        self.raw_data: list[Sample] = []
        self.time_point_1: float | str = "NaN"
        self.time_point_2: float | str = "NaN"
        self.measured_time: float | str = "NaN"
        self._start_monotonic = 0.0
        self._collecting = False
        self._selecting_point: int | None = None
        self._images: dict[str, ImageTk.PhotoImage] = {}

        self._build_header()
        self._build_plot_area()
        self._build_controls()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _set_windows_identity(self) -> None:
        if sys.platform != "win32":
            return
        try:
            windll.shell32.SetCurrentProcessExplicitAppUserModelID("Fairfield.CSBL.Agility")
        except Exception:
            pass

    def _set_app_icon(self) -> None:
        icon_path = resource_path("pythoncode", "assets", "lablogo.ico")
        if not icon_path.exists():
            icon_path = Path(__file__).resolve().parent / "assets" / "lablogo.ico"
        if icon_path.exists():
            try:
                self.iconbitmap(default=str(icon_path))
            except tk.TclError:
                pass

    def _build_header(self) -> None:
        header = tk.Frame(self, bg=FAIRFIELD_BLACK)
        header.place(relx=0, rely=0, relwidth=1, relheight=0.145)

        accent = tk.Frame(self, bg=FAIRFIELD_RED)
        accent.place(relx=0, rely=0.145, relwidth=1, height=5)

        lab_logo = self._load_image(resource_path("pics", "lablogo.png"), (92, 82))
        if lab_logo:
            tk.Label(header, image=lab_logo, bg=FAIRFIELD_BLACK).place(
                relx=0.022, rely=0.1, relwidth=0.08, relheight=0.78
            )

        title_block = tk.Frame(header, bg=FAIRFIELD_BLACK)
        title_block.place(relx=0.12, rely=0.16, relwidth=0.56, relheight=0.68)
        tk.Label(
            title_block,
            text=APP_NAME,
            bg=FAIRFIELD_BLACK,
            fg="white",
            font=("Segoe UI", 28, "bold"),
            anchor="w",
        ).pack(fill="x")
        tk.Label(
            title_block,
            text="Community-Situated Biomechanics Lab at Fairfield University",
            bg=FAIRFIELD_BLACK,
            fg=FAIRFIELD_GRAY,
            font=("Segoe UI", 12),
            anchor="w",
        ).pack(fill="x")

        fairfield_logo = self._load_image(resource_path("pics", "fairfieldlogo.png"), (94, 82))
        if fairfield_logo:
            tk.Label(header, image=fairfield_logo, bg=FAIRFIELD_BLACK).place(
                relx=0.895, rely=0.1, relwidth=0.075, relheight=0.78
            )

    def _load_image(self, path: Path, max_size: tuple[int, int]) -> ImageTk.PhotoImage | None:
        if not path.exists():
            return None
        image = Image.open(path).convert("RGBA")
        image.thumbnail(max_size, Image.LANCZOS)
        photo = ImageTk.PhotoImage(image)
        self._images[str(path)] = photo
        return photo

    def _build_plot_area(self) -> None:
        self.plot = tk.Canvas(self, bg=PANEL_BG, highlightthickness=1, highlightbackground=FAIRFIELD_GRAY)
        self.plot.place(relx=0.057478, rely=0.34, relwidth=0.880938, relheight=0.58)
        self.plot.bind("<Button-1>", self._plot_clicked)
        self.plot.bind("<Configure>", lambda _event: self._redraw_plot())

    def _build_controls(self) -> None:
        self.initialize_button = tk.Button(
            self, text="Initialize Arduino", command=self.initialize_arduino
        )
        self._style_button(self.initialize_button, primary=True)
        self.initialize_button.place(relx=0.028169, rely=0.18, relwidth=0.117371, relheight=0.065)

        port_panel = self._panel("Arduino Com Port", 0.166667, 0.18, 0.130282, 0.065)
        self.port_edit = tk.Entry(port_panel, justify="center")
        self._style_entry(self.port_edit)
        self.port_edit.insert(0, "com8")
        self.port_edit.place(relx=0.082822, rely=0.216495, relwidth=0.828221, relheight=0.703608)

        self.start_button = tk.Button(
            self, text="Start Data Collection", command=self.start_data_collection
        )
        self._style_button(self.start_button, primary=True)
        self.start_button.place(relx=0.318075, rely=0.18, relwidth=0.133803, relheight=0.065)

        duration_panel = self._panel("Sampling Duration", 0.473005, 0.18, 0.132629, 0.065)
        self.sampling_dur_edit = tk.Entry(duration_panel, justify="center")
        self._style_entry(self.sampling_dur_edit)
        self.sampling_dur_edit.insert(0, "10")
        self.sampling_dur_edit.place(relx=0.081571, rely=0.216495, relwidth=0.815710, relheight=0.703608)

        player_panel = self._panel("Enter Player Number", 0.626761, 0.18, 0.144366, 0.065)
        self.player_number_edit = tk.Entry(player_panel, justify="center")
        self._style_entry(self.player_number_edit)
        self.player_number_edit.insert(0, "Player Number")
        self.player_number_edit.place(relx=0.107772, rely=0.135309, relwidth=0.837306, relheight=0.703608)

        self.export_data_button = tk.Button(self, text="Export Data", command=self.export_data)
        self._style_button(self.export_data_button)
        self.export_data_button.place(relx=0.793427, rely=0.18, relwidth=0.177230, relheight=0.065)

        self.first_time_point_button = tk.Button(
            self, text="Select 1st Point", command=lambda: self._begin_selecting_point(1)
        )
        self._style_button(self.first_time_point_button)
        self.first_time_point_button.place(relx=0.028169, rely=0.255, relwidth=0.118545, relheight=0.055)

        time_1_panel = self._panel("1st Time Point", 0.201878, 0.255, 0.160798, 0.055)
        self.time_point_1_text = tk.Label(time_1_panel, text="No Data", bg=PANEL_BG, fg=TEXT_DARK)
        self.time_point_1_text.place(relx=0.052174, rely=0.215661, relwidth=0.916770, relheight=0.620026)

        self.second_time_point_button = tk.Button(
            self, text="Select 2nd Point", command=lambda: self._begin_selecting_point(2)
        )
        self._style_button(self.second_time_point_button)
        self.second_time_point_button.place(relx=0.417840, rely=0.255, relwidth=0.118545, relheight=0.055)

        time_2_panel = self._panel("2nd Time Point", 0.591549, 0.255, 0.160798, 0.055)
        self.time_point_2_text = tk.Label(time_2_panel, text="No Data", bg=PANEL_BG, fg=TEXT_DARK)
        self.time_point_2_text.place(relx=0.051852, rely=0.215661, relwidth=0.911111, relheight=0.673941)

        measured_panel = self._panel("Measured Time", 0.807512, 0.255, 0.160798, 0.055)
        self.measured_time_text = tk.Label(measured_panel, text="No Data", bg=PANEL_BG, fg=TEXT_DARK)
        self.measured_time_text.place(relx=0.029703, rely=0.215661, relwidth=0.913366, relheight=0.673941)

    def _panel(self, title: str, relx: float, rely: float, relwidth: float, relheight: float) -> tk.LabelFrame:
        panel = tk.LabelFrame(
            self,
            text=title,
            bg=PANEL_BG,
            fg=FAIRFIELD_BLACK,
            labelanchor="n",
            font=("Segoe UI", 9, "bold"),
            bd=1,
            relief="solid",
        )
        panel.place(relx=relx, rely=rely, relwidth=relwidth, relheight=relheight)
        return panel

    def _style_button(self, button: tk.Button, primary: bool = False) -> None:
        bg = FAIRFIELD_RED if primary else FAIRFIELD_BLACK
        active_bg = "#B90F28" if primary else "#3A3435"
        button.configure(
            bg=bg,
            fg="white",
            activebackground=active_bg,
            activeforeground="white",
            bd=0,
            relief="flat",
            cursor="hand2",
            font=("Segoe UI", 10, "bold"),
            highlightthickness=0,
            padx=6,
            pady=4,
        )

    def _style_entry(self, entry: tk.Entry) -> None:
        entry.configure(
            bg="#FFFFFF",
            fg=TEXT_DARK,
            insertbackground=FAIRFIELD_RED,
            relief="flat",
            font=("Segoe UI", 10),
        )

    def initialize_arduino(self) -> None:
        port = self.port_edit.get()
        try:
            self.arduino = ArduinoVoltageReader(port)
        except Exception as exc:
            self.port_edit.configure(bg="#FFD6DE", fg=ERROR_RED)
            messagebox.showerror("Arduino initialization failed", str(exc))
            return
        self.port_edit.configure(bg="#DDF3E6", fg=SUCCESS_GREEN)

    def start_data_collection(self) -> None:
        if self._collecting:
            return
        try:
            self.total_time = float(self.sampling_dur_edit.get())
        except ValueError:
            messagebox.showerror("Invalid duration", "Sampling Duration must be a number.")
            return
        if self.total_time <= 0:
            messagebox.showerror("Invalid duration", "Sampling Duration must be greater than zero.")
            return
        if self.arduino is None:
            messagebox.showerror("Arduino not initialized", "Click Initialize Arduino before collecting data.")
            return

        self.raw_data = []
        self.time_point_1 = "NaN"
        self.time_point_2 = "NaN"
        self.measured_time = "NaN"
        self._selecting_point = None
        self._set_status_label(self.time_point_1_text, "No Data", ERROR_RED)
        self._set_status_label(self.time_point_2_text, "No Data", ERROR_RED)
        self._set_status_label(self.measured_time_text, "No Data", ERROR_RED)
        self._collecting = True
        self._start_monotonic = time.monotonic()
        self._redraw_plot()
        self._collect_next_sample()

    def _collect_next_sample(self) -> None:
        elapsed = time.monotonic() - self._start_monotonic
        if elapsed >= self.total_time:
            self._collecting = False
            matlab_i = len(self.raw_data) + 1
            self.sample_rate = matlab_i / self.total_time
            return

        voltage = self.arduino.read_voltage_a1() if self.arduino else math.nan
        self.raw_data.append(Sample(elapsed, voltage))
        self._redraw_plot()
        self.after(1, self._collect_next_sample)

    def _begin_selecting_point(self, point_number: int) -> None:
        if not self.raw_data:
            messagebox.showwarning("No data", "Collect data before selecting points.")
            return
        self._selecting_point = point_number
        self.plot.configure(cursor="crosshair")

    def _plot_clicked(self, event: tk.Event) -> None:
        if self._selecting_point is None or not self.raw_data:
            return
        clicked_time = self._canvas_to_time(event.x)
        out_index = self._first_sample_after(clicked_time)
        sample = self.raw_data[out_index]

        if self._selecting_point == 1:
            self.time_point_1 = clicked_time
            self._set_status_label(self.time_point_1_text, self._matlab_num(clicked_time), SUCCESS_GREEN)
        else:
            self.time_point_2 = clicked_time
            if isinstance(self.time_point_1, float):
                self.measured_time = self.time_point_2 - self.time_point_1
            else:
                self.measured_time = "NaN"
            self._set_status_label(self.time_point_2_text, self._matlab_num(clicked_time), SUCCESS_GREEN)
            self._set_status_label(self.measured_time_text, self._matlab_num(self.measured_time), SUCCESS_GREEN)

        self._selecting_point = None
        self.plot.configure(cursor="")
        self._redraw_plot(selected=(sample, "Start Point" if self.time_point_2 == "NaN" else None))

    def _first_sample_after(self, time_s: float) -> int:
        for index, sample in enumerate(self.raw_data):
            if time_s < sample.time_s:
                return index
        return len(self.raw_data) - 1

    def _set_status_label(self, label: tk.Label, text: str, color: str) -> None:
        label.configure(text=text, bg=color, fg="white", font=("Segoe UI", 10, "bold"))

    def _matlab_num(self, value: float | str) -> str:
        if isinstance(value, str):
            return value
        return f"{value:.15g}"

    def export_data(self) -> None:
        if not self.raw_data:
            messagebox.showwarning("No data", "There is no raw data to export.")
            return

        export_name = f"agility_{self.player_number_edit.get()}.xlsx"
        export_dir = MATLAB_EXPORT_DIR if MATLAB_EXPORT_DIR.exists() else Path.cwd()
        output_path = export_dir / export_name

        now = datetime.now()
        test_info = [
            "Player Number",
            "year",
            "month",
            "day",
            "hour",
            "minute",
            "second",
            "Sample Duration",
            "Sample Rate",
            "Time 1",
            "Time 2",
            "Time Measured",
        ]
        test_results = [
            self.player_number_edit.get(),
            str(now.year),
            str(now.month),
            str(now.day),
            str(now.hour),
            str(now.minute),
            str(now.second),
            self.total_time,
            self.sample_rate,
            self.time_point_1,
            self.time_point_2,
            self.measured_time,
        ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        for row, (label, value) in enumerate(zip(test_info, test_results), start=1):
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=2, value=value)

        sheet.cell(row=2, column=5, value="Time (s)")
        sheet.cell(row=2, column=6, value="Signal (V)")
        for row, sample in enumerate(self.raw_data, start=3):
            sheet.cell(row=row, column=5, value=sample.time_s)
            sheet.cell(row=row, column=6, value=sample.voltage)

        workbook.save(output_path)
        messagebox.showinfo("Export Data", f"Saved {output_path}")

    def _redraw_plot(self, selected: tuple[Sample, str | None] | None = None) -> None:
        self.plot.delete("all")
        width = max(self.plot.winfo_width(), 2)
        height = max(self.plot.winfo_height(), 2)
        margin_left, margin_right, margin_top, margin_bottom = 52, 18, 18, 38
        x0, y0 = margin_left, height - margin_bottom
        x1, y1 = width - margin_right, margin_top

        self.plot.create_rectangle(x0, y1, x1, y0, outline=FAIRFIELD_BLACK, fill="white", width=2)
        self.plot.create_text((x0 + x1) / 2, height - 12, text="Time (s)", fill=TEXT_DARK)
        self.plot.create_text(16, (y0 + y1) / 2, text="Signal (V)", angle=90, fill=TEXT_DARK)

        for frac in (0, 0.25, 0.5, 0.75, 1):
            x = x0 + frac * (x1 - x0)
            y = y0 - frac * (y0 - y1)
            self.plot.create_line(x, y0, x, y1, fill="#ECEFF1")
            self.plot.create_line(x0, y, x1, y, fill="#ECEFF1")
            self.plot.create_text(x, y0 + 14, text=self._matlab_num(frac * self.total_time), fill=TEXT_DARK)
            self.plot.create_text(x0 - 18, y, text=self._matlab_num(frac * 5), fill=TEXT_DARK)

        if len(self.raw_data) == 1:
            x, y = self._to_canvas(self.raw_data[0], x0, y0, x1, y1)
            self.plot.create_oval(x - 3, y - 3, x + 3, y + 3, fill=FAIRFIELD_BLACK, outline=FAIRFIELD_BLACK)
        elif len(self.raw_data) > 1:
            points: list[float] = []
            for sample in self.raw_data:
                points.extend(self._to_canvas(sample, x0, y0, x1, y1))
            self.plot.create_line(*points, fill=FAIRFIELD_RED, width=2)

        self._draw_selected_point(self.time_point_1, "Start Point", x0, y0, x1, y1)
        self._draw_selected_point(self.time_point_2, "End Point", x0, y0, x1, y1)
        if selected:
            sample, label = selected
            self._draw_marker(sample, label, x0, y0, x1, y1)

    def _draw_selected_point(
        self, point: float | str, label: str, x0: int, y0: int, x1: int, y1: int
    ) -> None:
        if not isinstance(point, float) or not self.raw_data:
            return
        self._draw_marker(self.raw_data[self._first_sample_after(point)], label, x0, y0, x1, y1)

    def _draw_marker(
        self, sample: Sample, label: str | None, x0: int, y0: int, x1: int, y1: int
    ) -> None:
        x, y = self._to_canvas(sample, x0, y0, x1, y1)
        self.plot.create_oval(x - 5, y - 5, x + 5, y + 5, fill=FAIRFIELD_BLACK, outline=FAIRFIELD_RED, width=2)
        if label:
            self.plot.create_text(x + 55, y - 10, text=label, fill=FAIRFIELD_BLACK, font=("Segoe UI", 10, "bold"))

    def _to_canvas(self, sample: Sample, x0: int, y0: int, x1: int, y1: int) -> tuple[float, float]:
        x = x0 + (sample.time_s / max(self.total_time, 1e-9)) * (x1 - x0)
        voltage = 0.0 if math.isnan(sample.voltage) else max(0.0, min(5.0, sample.voltage))
        y = y0 - (voltage / 5.0) * (y0 - y1)
        return x, y

    def _canvas_to_time(self, x: int) -> float:
        width = max(self.plot.winfo_width(), 2)
        margin_left, margin_right = 52, 18
        x0, x1 = margin_left, width - margin_right
        frac = (x - x0) / max(x1 - x0, 1)
        return max(0.0, min(self.total_time, frac * self.total_time))

    def _on_close(self) -> None:
        if self.arduino is not None:
            self.arduino.close()
        self.destroy()


def main() -> int:
    app = AgilityGui()
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
